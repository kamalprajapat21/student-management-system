from fastapi import APIRouter, Depends, Response
from app.utils.auth_deps import get_current_user, require_admin, require_teacher_or_admin
from app.utils.helpers import serialize_list
from app.database import get_db
from bson import ObjectId
from datetime import datetime, timedelta
import io, json

router = APIRouter(prefix="/api/analytics", tags=["Analytics"])


@router.get("/student/{student_id}")
async def student_analytics(student_id: str, current_user=Depends(get_current_user)):
    db = get_db()
    # Attendance summary
    records = await db.attendance.find({"student_id": student_id}).to_list(1000)
    total = len(records)
    present = sum(1 for r in records if r["status"] in ["present", "late"])
    attendance_pct = round((present / total * 100), 2) if total else 0

    # Subject-wise attendance
    subj_pipeline = [
        {"$match": {"student_id": student_id}},
        {"$group": {
            "_id": "$subject",
            "total": {"$sum": 1},
            "present": {"$sum": {"$cond": [{"$in": ["$status", ["present", "late"]]}, 1, 0]}}
        }}
    ]
    subj_attendance = await db.attendance.aggregate(subj_pipeline).to_list(20)

    # Marks summary
    marks = await db.marks.find({"student_id": student_id}).to_list(100)
    marks_by_subject = {}
    for m in marks:
        exam = await db.exams.find_one({"_id": ObjectId(m["exam_id"])})
        if exam:
            subj = exam["subject"]
            if subj not in marks_by_subject:
                marks_by_subject[subj] = {"obtained": 0, "total": 0, "count": 0}
            marks_by_subject[subj]["obtained"] += m.get("marks_obtained", 0)
            marks_by_subject[subj]["total"] += exam.get("total_marks", 100)
            marks_by_subject[subj]["count"] += 1

    # Assignment completion
    assignments = await db.assignments.find(
        {"class_id": current_user.get("class_name")}
    ).to_list(100)
    total_assignments = len(assignments)
    submitted = 0
    for a in assignments:
        sub = await db.submissions.find_one({"assignment_id": str(a["_id"]), "student_id": student_id})
        if sub:
            submitted += 1

    return {
        "attendance": {
            "total": total,
            "present": present,
            "percentage": attendance_pct,
            "subject_wise": subj_attendance
        },
        "marks": {
            "by_subject": marks_by_subject
        },
        "assignments": {
            "total": total_assignments,
            "submitted": submitted,
            "completion_rate": round(submitted / total_assignments * 100, 2) if total_assignments else 0
        }
    }


@router.get("/teacher/class/{class_id}")
async def teacher_class_analytics(class_id: str, current_user=Depends(require_teacher_or_admin)):
    db = get_db()
    students = await db.users.find(
        {"class_name": class_id, "role": "student", "is_active": True}, {"password": 0}
    ).to_list(200)

    student_analytics = []
    for student in students:
        sid = str(student["_id"])
        records = await db.attendance.find({"student_id": sid}).to_list(500)
        total = len(records)
        present = sum(1 for r in records if r["status"] in ["present", "late"])
        pct = round(present / total * 100, 2) if total else 0

        marks = await db.marks.find({"student_id": sid}).to_list(50)
        avg_marks = 0
        if marks:
            total_pct = sum(
                (m["marks_obtained"] / 100 * 100) for m in marks if m.get("marks_obtained") is not None
            )
            avg_marks = round(total_pct / len(marks), 2)

        student_analytics.append({
            "student_id": sid,
            "name": student["full_name"],
            "roll_number": student.get("roll_number"),
            "attendance_percentage": pct,
            "average_marks": avg_marks,
            "risk": "high" if pct < 60 or avg_marks < 40 else "medium" if pct < 75 or avg_marks < 60 else "low"
        })

    top_performers = sorted(student_analytics, key=lambda x: x["average_marks"], reverse=True)[:5]
    weak_students = [s for s in student_analytics if s["risk"] == "high"]

    return {
        "total_students": len(students),
        "class_analytics": student_analytics,
        "top_performers": top_performers,
        "weak_students": weak_students,
        "avg_attendance": round(sum(s["attendance_percentage"] for s in student_analytics) / len(student_analytics), 2) if student_analytics else 0
    }


@router.get("/admin/overview")
async def admin_overview(current_user=Depends(require_admin)):
    db = get_db()
    total_students = await db.users.count_documents({"role": "student", "is_active": True})
    total_teachers = await db.users.count_documents({"role": "teacher", "is_active": True})
    total_parents = await db.users.count_documents({"role": "parent", "is_active": True})

    # Attendance this week
    week_start = (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%d")
    total_records = await db.attendance.count_documents({"date": {"$gte": week_start}})
    present_records = await db.attendance.count_documents({
        "date": {"$gte": week_start},
        "status": {"$in": ["present", "late"]}
    })

    # Fee stats
    fee_pipeline = [
        {"$group": {
            "_id": None,
            "total_collected": {"$sum": "$amount_paid"},
            "total_due": {"$sum": "$due_amount"}
        }}
    ]
    fee_stats = await db.fees.aggregate(fee_pipeline).to_list(1)

    # Students by department
    dept_pipeline = [
        {"$match": {"role": "student", "is_active": True}},
        {"$group": {"_id": "$department", "count": {"$sum": 1}}}
    ]
    by_dept = await db.users.aggregate(dept_pipeline).to_list(20)

    return {
        "total_students": total_students,
        "total_teachers": total_teachers,
        "total_parents": total_parents,
        "weekly_attendance": {
            "total": total_records,
            "present": present_records,
            "percentage": round(present_records / total_records * 100, 2) if total_records else 0
        },
        "fee_stats": fee_stats[0] if fee_stats else {"total_collected": 0, "total_due": 0},
        "students_by_department": by_dept
    }


@router.get("/admin/export/students")
async def export_students_excel(current_user=Depends(require_admin)):
    db = get_db()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    students = await db.users.find({"role": "student", "is_active": True}, {"password": 0}).to_list(1000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["Roll Number", "Full Name", "Email", "Department", "Class", "Semester", "Phone", "Joined"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.get("roll_number", ""))
        ws.cell(row=row, column=2, value=student.get("full_name", ""))
        ws.cell(row=row, column=3, value=student.get("email", ""))
        ws.cell(row=row, column=4, value=student.get("department", ""))
        ws.cell(row=row, column=5, value=student.get("class_name", ""))
        ws.cell(row=row, column=6, value=student.get("semester", ""))
        ws.cell(row=row, column=7, value=student.get("phone", ""))
        ws.cell(row=row, column=8, value=str(student.get("created_at", "")))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"}
    )
